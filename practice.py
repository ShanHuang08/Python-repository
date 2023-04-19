<<<<<<< HEAD

def For_Bingo():
    ans=45
    guess_count=0
    guess_limit=5

    for guess in range(guess_limit):
        guess_count+=1
        guess=int(input("Innput digit:"))
        #Bingo跟次數無關 可以分開寫
        if guess == ans:
            print("Bingo!")
            break
        if guess_count < guess_limit:
            if guess < ans:
                print("Lower")
            elif guess > ans:
                print("Larger") 
        else:
            print("GG")

def While_Bingo():
    ans=35
    guess=None
    guess_count=0
    guess_limit=5
    out_of_limit=False #一開始沒有超過limit 所以設false

    while ans != guess and not(out_of_limit):
        guess_count+=1
        print("Total: "+str(guess_count)+" times")
        if guess_count <= guess_limit:
            guess=int(input("猜一個數字:"))
            if guess < ans:
                print("猜錯了,數字太小")
            elif guess > ans:
                print("猜錯了,數字太大")
        else:
            out_of_limit=True

    if out_of_limit and guess != ans:
        print("GG")
    else:
        print("Bingo!")

import random
def scissors():   
    guess=None
    scissors="剪刀"
    rock="石頭"
    paper="布"
    Result=[scissors,rock,paper]

    while True:
        guess=input("出拳: ")
        answer=random.choice(Result)  
        if guess == answer:
            print("平手")
        elif guess == scissors and answer == rock:
            print("我出"+str(answer)+", 你輸了")
        elif guess == scissors and answer == paper:
            print("我出"+str(answer)+", 你贏了")
        elif guess == rock and answer == paper:
            print("我出"+str(answer)+", 你輸了")
        elif guess == rock and answer == scissors:
            print("我出"+str(answer)+", 你贏了")
        elif guess == paper and answer == scissors:
            print("我出"+str(answer)+", 你輸了")
        elif guess == paper and answer == rock:
            print("我出"+str(answer)+", 你贏了")
        else:
            print("輸入錯誤, 重來")

def keywordtest():
    i=1
    keyword_list=[]
    while i<=3:
        keyword_input=input('第'+str(i)+'個關鍵字= ')
        keyword_list.append(keyword_input)
        i+=1

    print(keyword_list)

def titlelist(ele):
    Title = ['日期','新增確診數','總確診數','新增確診數/百萬','新聞稿發佈新增確診數' ]
    Title.append(ele)
    print(Title)

def monday():
    n=input('輸入數字(1-7): ')
    sevendays='一二三四五六日'
    if n.isdigit()!=True:
        print('輸入錯誤')
    elif int(n)<=0 or int(n)>7:
        print('請輸入正確數字範圍')
    elif n in ['1','2','3','4','5','6','7']:
        print('星期'+sevendays[int(n)-1])
    else:
        print('輸入錯誤')

def month():
    months='JanFebMarAprMayJunJulAugSepOctNovDec'
    n=input('輸入月份(1-12): ')
    pos=(int(n)-1)*3
    print('pos='+str(pos))
    print(months[pos:pos+3]+'.')

from math import sqrt
from time import process_time
from random import randint, uniform
def Pi():
    Dots=2**20
    Hits=0
    diameter=3
    cost=process_time()
    for i in range(0,Dots):
        x,y=uniform(0,diameter),uniform(0,diameter)
        distance=sqrt(x**2+y**2)
        if distance<=diameter:
            Hits+=1
    pi=4*(Hits/Dots)
    print(f'Pi= {pi}')
    print(f'Process time= {cost}s')

def TryExcept():
    try:   
        num1,num2=eval(input('Enter 2 numbers ex:1,2 '))
        result=num1/num2
    except ZeroDivisionError:
        print('Division by zero')
    except SyntaxError:
        print('A comma(,) may be missing in the input')
    except:
        print('Something went wrong while input')
    else:
        print(f'No exception, result={result}')
    finally:
        print('Executing the final clause')    
    
def FindMaxNum():
    AmountNum=eval(input('How many numbers will generate? '))
    # max=eval(input('Set a init max number= '))
    max=AmountNum
    
    for i in range (1,AmountNum+1):
        Nums=randint(0,max*2)
        # print(f'{i}. {Nums}')
        if Nums>max:
            max=Nums
            print(f'{i}. {max} is larger')
    print(f'Through {AmountNum} times comparisons, the largest value is: {max}')
        
def eval():
    a,b,c=eval(input('input a: ')),eval(input('input b: ')),eval(input('input c: '))
    print(f'{a}+{b}+{c}={a+b+c}')

def PrimeNumber():
    max=10
    for n in range (2,max+1):
        for j in range(2,n):
            if n%j==0:
                break
        else:
            print(f'{n} is a prime number')

def PrimeNumber2():
    num=2
    max=10
    while num<=max:
        i=2
        b=True
        while i<num:
            if num%i==0:
                b=False
            i+=1
        if b:
            print(f'{num} is a prime number')
        # else:
        #     print(f'{num} is not a prime number')
        num+=1
        




def test():
    a=12345
    List=[]
    for i in range(len(str(a))-1,-1,-1):
        List.append(str(a)[i])
    res=''.join(List)
    print(res)

    b='abcde'
    List2=[]
    for k in range(len(b)-1,-1,-1):
        List2.append(b[k])
    res2=''.join(List2)
    print(res2)

    min=1
    max=10
    sum=0
    while min<=max:
        sum=sum+min
        min+=1
    print(f'sum1= {sum}')

    for j in range(min,max+1):
        sum=sum+min
    print(f'sum2= {sum}')

from openpyxl import Workbook
def CreateSheet():
    wb=Workbook()
    ws1=wb.active
    ws1.title='sheet1'
    wb.create_sheet('sheet2')
    wb.create_sheet('sheet3')
    wb.create_sheet('sheet4')
    print(wb.sheetnames)

    ws1=wb['sheet1']
    ws1.append(['工作表100','工作表101'])
    ws1.append(['工作表110','工作表111'])

    ws2=wb['sheet2']
    ws2.append(['工作表200','工作表201'])
    ws2.append(['工作表210','工作表211'])

    ws3=wb['sheet3']
    ws3.append(['工作表300','工作表301'])
    ws3.append(['工作表310','工作表311'])

    wb.save('test.xlsx')


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC 
def Element_isclickable(xpath):
    driver = webdriver.Chrome()
    driver.get('https://www.bing.com/')
    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, xpath)))
        return True
    except :
        return False


def functions():
    For_Bingo()
    While_Bingo()
    titlelist('abc')
    monday()
    month()
    test()
    Pi()
    TryExcept()
    FindMaxNum()
    eval()
    PrimeNumber()
    PrimeNumber2()
    scissors()
    CreateSheet()
    Element_isclickable(xpath)
While_Bingo()

=======

def For_Bingo():
    ans=45
    guess_count=0
    guess_limit=5

    for guess in range(guess_limit):
        guess_count+=1
        guess=int(input("Innput digit:"))
        #Bingo跟次數無關 可以分開寫
        if guess == ans:
            print("Bingo!")
            break
        if guess_count < guess_limit:
            if guess < ans:
                print("Lower")
            elif guess > ans:
                print("Larger") 
        else:
            print("GG")

def While_Bingo():
    ans=35
    guess=None
    guess_count=0
    guess_limit=5
    out_of_limit=False #一開始沒有超過limit 所以設false

    while ans != guess and not(out_of_limit):
        guess_count+=1
        print("Total: "+str(guess_count)+" times")
        if guess_count <= guess_limit:
            guess=int(input("猜一個數字:"))
            if guess < ans:
                print("猜錯了,數字太小")
            elif guess > ans:
                print("猜錯了,數字太大")
        else:
            out_of_limit=True

    if out_of_limit and guess != ans:
        print("GG")
    else:
        print("Bingo!")

import random
def scissors():   
    guess=None
    scissors="剪刀"
    rock="石頭"
    paper="布"
    Result=[scissors,rock,paper]

    while True:
        guess=input("出拳: ")
        answer=random.choice(Result)  
        if guess == answer:
            print("平手")
        elif guess == scissors and answer == rock:
            print("我出"+str(answer)+", 你輸了")
        elif guess == scissors and answer == paper:
            print("我出"+str(answer)+", 你贏了")
        elif guess == rock and answer == paper:
            print("我出"+str(answer)+", 你輸了")
        elif guess == rock and answer == scissors:
            print("我出"+str(answer)+", 你贏了")
        elif guess == paper and answer == scissors:
            print("我出"+str(answer)+", 你輸了")
        elif guess == paper and answer == rock:
            print("我出"+str(answer)+", 你贏了")
        else:
            print("輸入錯誤, 重來")

def keywordtest():
    i=1
    keyword_list=[]
    while i<=3:
        keyword_input=input('第'+str(i)+'個關鍵字= ')
        keyword_list.append(keyword_input)
        i+=1

    print(keyword_list)

def titlelist(ele):
    Title = ['日期','新增確診數','總確診數','新增確診數/百萬','新聞稿發佈新增確診數' ]
    Title.append(ele)
    print(Title)

def monday():
    n=input('輸入數字(1-7): ')
    sevendays='一二三四五六日'
    if n.isdigit()!=True:
        print('輸入錯誤')
    elif int(n)<=0 or int(n)>7:
        print('請輸入正確數字範圍')
    elif n in ['1','2','3','4','5','6','7']:
        print('星期'+sevendays[int(n)-1])
    else:
        print('輸入錯誤')

def month():
    months='JanFebMarAprMayJunJulAugSepOctNovDec'
    n=input('輸入月份(1-12): ')
    pos=(int(n)-1)*3
    print('pos='+str(pos))
    print(months[pos:pos+3]+'.')

from math import sqrt
from time import process_time
from random import randint, uniform
def Pi():
    Dots=2**20
    Hits=0
    diameter=3
    cost=process_time()
    for i in range(0,Dots):
        x,y=uniform(0,diameter),uniform(0,diameter)
        distance=sqrt(x**2+y**2)
        if distance<=diameter:
            Hits+=1
    pi=4*(Hits/Dots)
    print(f'Pi= {pi}')
    print(f'Process time= {cost}s')

def TryExcept():
    try:   
        num1,num2=eval(input('Enter 2 numbers ex:1,2 '))
        result=num1/num2
    except ZeroDivisionError:
        print('Division by zero')
    except SyntaxError:
        print('A comma(,) may be missing in the input')
    except:
        print('Something went wrong while input')
    else:
        print(f'No exception, result={result}')
    finally:
        print('Executing the final clause')    
    
def FindMaxNum():
    AmountNum=eval(input('How many numbers will generate? '))
    # max=eval(input('Set a init max number= '))
    max=AmountNum
    
    for i in range (1,AmountNum+1):
        Nums=randint(0,max*2)
        # print(f'{i}. {Nums}')
        if Nums>max:
            max=Nums
            print(f'{i}. {max} is larger')
    print(f'Through {AmountNum} times comparisons, the largest value is: {max}')
        
def eval():
    a,b,c=eval(input('input a: ')),eval(input('input b: ')),eval(input('input c: '))
    print(f'{a}+{b}+{c}={a+b+c}')

def PrimeNumber():
    max=10
    for n in range (2,max+1):
        for j in range(2,n):
            if n%j==0:
                break
        else:
            print(f'{n} is a prime number')

def PrimeNumber2():
    num=2
    max=10
    while num<=max:
        i=2
        b=True
        while i<num:
            if num%i==0:
                b=False
            i+=1
        if b:
            print(f'{num} is a prime number')
        # else:
        #     print(f'{num} is not a prime number')
        num+=1
        




def test():
    a=12345
    List=[]
    for i in range(len(str(a))-1,-1,-1):
        List.append(str(a)[i])
    res=''.join(List)
    print(res)

    b='abcde'
    List2=[]
    for k in range(len(b)-1,-1,-1):
        List2.append(b[k])
    res2=''.join(List2)
    print(res2)

    min=1
    max=10
    sum=0
    while min<=max:
        sum=sum+min
        min+=1
    print(f'sum1= {sum}')

    for j in range(min,max+1):
        sum=sum+min
    print(f'sum2= {sum}')

from openpyxl import Workbook
def CreateSheet():
    wb=Workbook()
    ws1=wb.active
    ws1.title='sheet1'
    wb.create_sheet('sheet2')
    wb.create_sheet('sheet3')
    wb.create_sheet('sheet4')
    print(wb.sheetnames)

    ws1=wb['sheet1']
    ws1.append(['工作表100','工作表101'])
    ws1.append(['工作表110','工作表111'])

    ws2=wb['sheet2']
    ws2.append(['工作表200','工作表201'])
    ws2.append(['工作表210','工作表211'])

    ws3=wb['sheet3']
    ws3.append(['工作表300','工作表301'])
    ws3.append(['工作表310','工作表311'])

    wb.save('test.xlsx')


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC 
def Element_isclickable(xpath):
    driver = webdriver.Chrome()
    driver.get('https://www.bing.com/')
    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, xpath)))
        return True
    except :
        return False


def functions():
    For_Bingo()
    While_Bingo()
    titlelist('abc')
    monday()
    month()
    test()
    Pi()
    TryExcept()
    FindMaxNum()
    eval()
    PrimeNumber()
    PrimeNumber2()
    scissors()
    CreateSheet()
    Element_isclickable(xpath)
While_Bingo()

>>>>>>> 8ed83d2505d4355229a31d2a06788e31225895b9
