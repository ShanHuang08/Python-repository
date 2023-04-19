from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook
import time

# Crash.Net scrape
def CrashNet(wb,browser):
    ws=wb.active
    ws.title='Crash.Net'
    ws.append(['Crash.Net MotoGP News'])
    ws.append(['時間','連結','標題'])

    browser.get('https://www.crash.net/motogp/news')
    page=1
    while page<=2:   
        CreatedTime=browser.find_elements(By.XPATH,value='//div[@class="views-field views-field-created"]/span[@class="field-content"]')
        Titles=browser.find_elements(By.XPATH,value='//*[@id="block-system-main"]/div/div/div[1]/div/div[@class="views-field views-field-title"]/span/a')

        Get_CreatedTime=[]
        for times in CreatedTime:
            Get_CreatedTime.append(times.text)

        Get_Links=[]
        Get_Titles=[]
        for i in Titles:
            Links=i.get_attribute('href')
            Get_Links.append(Links)
            Get_Titles.append(i.text)

        for i in range(len(Get_CreatedTime)):
            Result=[]
            Result.append(Get_CreatedTime[i])
            Result.append(Get_Links[i])
            Result.append(Get_Titles[i])
            ws.append(Result)
        if page<2:
            browser.get('https://www.crash.net/motogp/news?page=1')
        page+=1
    print(f'CrashNet OK')

# Motorsport
def Motorsport(wb,browser):
    wb.create_sheet('Motorsport')
    ws2=wb['Motorsport']
    ws2.append(['Motorsport MotoGP News'])
    ws2.append(['時間','連結','標題'])

    browser.get('https://www.motorsport.com/motogp/news/')
    CreatedTime2=browser.find_elements(By.XPATH,value='//time[@class="ms-item_date-value"]')
    Title2_1st=browser.find_elements(By.XPATH,value='//*[@id="app_article_browse"]/div[4]/div[3]/div[3]/div/div[1]/div/div[3]/div[@class="ms-item--art "]/div[@class="ms-item_info"]/p/a')
    Titles2_2nd=browser.find_elements(By.XPATH,value='//*[@id="app_article_browse"]/div[4]/div[3]/div[3]/div/div[1]/div/div[1]/div[@class="ms-item--art "]/div[@class="ms-item_info"]/p/a')

    Get_CreatedTime2=[]
    for times in CreatedTime2:
        Time=times.get_attribute('datetime')
        Get_CreatedTime2.append(Time)
    # print(len(Get_CreatedTime2))

    Get_Links2=[]
    for link in Title2_1st:
        Link=link.get_attribute('href')
        Get_Links2.append(Link)
    for link in Titles2_2nd:
        Link=link.get_attribute('href')
        Get_Links2.append(Link)
    # print(len(Get_Links2))

    Get_Titles2=[]
    for title in Title2_1st:
        Get_Titles2.append(title.text)
    for title in Titles2_2nd:
        Get_Titles2.append(title.text)
    # print(len(Get_Titles2))

    for i in range(len(Get_CreatedTime2)):
        Result=[]
        Result.append(Get_CreatedTime2[i])
        Result.append(Get_Links2[i])
        Result.append(Get_Titles2[i])
        ws2.append(Result)
    print(f'Motorsport OK')

# TheRace
def TheRace(wb,browser):
    wb.create_sheet('TheRace')
    ws3=wb['TheRace']
    ws3.append(['The-Race News'])
    ws3.append(['連結','標題'])

    browser.get('https://the-race.com/category/motogp/')
    browser.maximize_window()
    element = WebDriverWait(browser, 2).until(
            EC.presence_of_element_located(((By.ID,'qc-cmp2-ui'))
        ))
    browser.find_element(By.XPATH,value='//*[@id="qc-cmp2-ui"]/div[2]/div/button[2]').click()
    time.sleep(3)
    browser.find_element(By.XPATH,value='//*[@id="TheRaceBody"]/div[4]/div[1]/div[1]/button[1]/img').click()
    body=browser.find_element(By.XPATH,value='/html/body').click()
    for i in range(3):
        ActionChains(browser).send_keys(Keys.PAGE_DOWN).perform()

    for i in range(5):
        browser.find_element(By.XPATH,value='//*[@id="therace_related_a"]/div/div[2]/div[2]/span/a').click()
        time.sleep(4)
        if browser.find_element(By.XPATH,value='//*[@id="therace_related_a"]/div/div[2]/div[2]/span/a').is_displayed()==False:
            ActionChains(browser).send_keys(Keys.PAGE_DOWN).perform()

    H1Title=browser.find_elements(By.XPATH,value='//*[@id="content"]/div/div[1]/div[2]/div[2]/h1/a')
    H2Title=browser.find_elements(By.XPATH,value='//*[@id="content"]/div/div[2]/div/div[@class="post-group"]/div/h2/a')
    H3Title_1st=browser.find_elements(By.XPATH,value='//*[@id="therace_related_b"]/div/div/div[@class="col-md-4 related_group"]/h3/a')
    H3Title_2nd=browser.find_elements(By.XPATH,value='//*[@id="therace_related_a"]/div/div[1]/div[@class="col-md-4 related_group"]/h3/a')

    Get_Linsk3=[]
    Get_Titles3=[]

    for i in H1Title:
        Link=i.get_attribute('href')
        Get_Linsk3.append(Link)
        Get_Titles3.append(i.text)
    for i in H2Title:
        Link=i.get_attribute('href')
        Get_Linsk3.append(Link)
        Get_Titles3.append(i.text)
    for i in H3Title_1st:
        Link=i.get_attribute('href')
        Get_Linsk3.append(Link)
        Get_Titles3.append(i.text)
    for i in H3Title_2nd:
        Link=i.get_attribute('href')
        Get_Linsk3.append(Link)
        Get_Titles3.append(i.text)
    for i in range(1,len(Get_Titles3)):
        Data=Get_Titles3[i].lower()
        Get_Titles3.pop(i)
        Get_Titles3.insert(i,Data)

    for i in range(len(Get_Titles3)):
        Result=[]
        Result.append(Get_Linsk3[i])
        Result.append(Get_Titles3[i])
        ws3.append(Result)
    print(f'TheRace OK')

def main():
    wb=Workbook()
    options=Options()
    options.add_argument('--headless')
    # FLoC permissions-policy: interest-cohort=().
    options.add_argument('--log-level=1')
    options.add_argument('--log-level=2')
    browser=webdriver.Chrome('./chromedriver.exe',options=options)
    # Websites
    CrashNet(wb,browser)
    Motorsport(wb,browser)
    TheRace(wb,browser)
    wb.save('MotoGP.xlsx')
    browser.quit()
main()